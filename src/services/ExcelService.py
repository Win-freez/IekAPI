import logging
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from pydantic import ValidationError

from src.services.IekService import IekAPI
from src.schemas.product import ProductResponse

logger = logging.getLogger(__name__)


class ExcelPriceCheckerOpenpyxl:
    PRICE_FIELDS = ["priceBase", "pricePersonal", "priceRoc", "priceRrc"]

    def __init__(self, api: IekAPI):
        self.api = api

    async def process_excel_file(
        self,
        file_path: str,
        output_path: str,
        article_column: str = "Article",
        header_row: int = 1,
    ) -> str:
        """
        Открывает существующий Excel file_path, добавляет справа 4 колонки с ценами
        и заполняет их значениями из IekAPI.get_products.
        Если значение цены отсутствует / None / не приводимо -> записывается "ОШИБКА".
        Возвращает путь сохранённого файла (output_path).
        """
        wb = load_workbook(filename=file_path)
        ws = wb.active

        # Найти колонку с артикулами по заголовку
        article_col_idx: Optional[int] = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col).value == article_column:
                article_col_idx = col
                break
        if article_col_idx is None:
            raise ValueError(f"Колонка с заголовком '{article_column}' не найдена в строке {header_row}")

        # Собираем строки по артикулам (включая пустые строки — будут помечены ОШИБКА)
        rows_by_article: Dict[str, List[int]] = {}
        for row in range(header_row + 1, ws.max_row + 1):
            raw = ws.cell(row=row, column=article_col_idx).value
            art = str(raw).strip() if raw is not None and str(raw).strip() != "" else ""
            rows_by_article.setdefault(art, []).append(row)

        # Подготовим колонки для цен — добавим справа (чтобы точно были после всех существующих)
        start_col = ws.max_column + 1
        price_col_indexes: List[int] = []
        for i, fld in enumerate(self.PRICE_FIELDS):
            col_idx = start_col + i
            ws.cell(row=header_row, column=col_idx, value=fld)
            price_col_indexes.append(col_idx)

        # Запрос к API
        unique_articles = list(rows_by_article.keys())
        products_results = await self.api.get_products(unique_articles)

        # Составляем словарь article -> raw result
        result_map: Dict[str, Any] = {}
        for item in products_results:
            art = item.get("article")
            key = str(art).strip() if art is not None else ""
            result_map[key] = item.get("result")

        # Вспомогательные функции записи
        def write_error_to_row(r: int):
            for col in price_col_indexes:
                ws.cell(row=r, column=col, value="ОШИБКА")

        def write_value_to_cell(r: int, col_index: int, value: Any):
            ws.cell(row=r, column=col_index, value=value)

        def safe_to_float(v: Any) -> Optional[float]:
            if v is None:
                return None
            try:
                return float(v)
            except Exception:
                return None

        # Основной проход: для каждого артикула — для каждой строки записываем 4 ячейки
        for art, rows in rows_by_article.items():
            res = result_map.get(art)

            # Если результата нет или это Exception — помечаем ОШИБКА во всех полях
            if res is None or isinstance(res, Exception):
                for r in rows:
                    write_error_to_row(r)
                continue

            # Если уже экземпляр ProductResponse
            if isinstance(res, ProductResponse):
                for r in rows:
                    for i, fld in enumerate(self.PRICE_FIELDS):
                        val = getattr(res, fld, None)
                        fv = safe_to_float(val)
                        if fv is None:
                            write_value_to_cell(r, price_col_indexes[i], "ОШИБКА")
                        else:
                            write_value_to_cell(r, price_col_indexes[i], fv)
                continue

            # Если dict-like — попробуем валидировать через Pydantic (модель допускает None в полях)
            if isinstance(res, dict):
                try:
                    product = ProductResponse.model_validate(res)
                    # используем модель (она может содержать None в полях)
                    for r in rows:
                        for i, fld in enumerate(self.PRICE_FIELDS):
                            val = getattr(product, fld, None)
                            fv = safe_to_float(val)
                            if fv is None:
                                write_value_to_cell(r, price_col_indexes[i], "ОШИБКА")
                            else:
                                write_value_to_cell(r, price_col_indexes[i], fv)
                    continue
                except ValidationError as ve:
                    # маловероятно, но если валидация не удалась — делаем максимально терпимое извлечение
                    logger.warning("ValidationError для артикула %s: %s", art, ve)
                    for r in rows:
                        for i, fld in enumerate(self.PRICE_FIELDS):
                            raw_val = res.get(fld) if isinstance(res, dict) else None
                            fv = safe_to_float(raw_val)
                            if fv is None:
                                write_value_to_cell(r, price_col_indexes[i], "ОШИБКА")
                            else:
                                write_value_to_cell(r, price_col_indexes[i], fv)
                    continue
                except Exception as e:
                    logger.exception("Unexpected error validating article %s: %s", art, e)
                    for r in rows:
                        write_error_to_row(r)
                    continue

            # Неподдерживаемый формат результата — помечаем как ОШИБКА
            for r in rows:
                write_error_to_row(r)

        # Сохраняем результат
        wb.save(output_path)
        logger.info("Результат сохранён в %s", output_path)
        return output_path
